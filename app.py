import streamlit as st
import pandas as pd
import requests
import io
import json
import re
import time
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials
import traceback

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
UC_BASE      = "http://129.154.230.19/unicommerce"   # Internal IP for all calls
UC_USERNAME  = "yashasavi@headphonezone.in"
UC_FACILITY  = "Warehouse"
GODOWN       = "Chennai Wh -Good"
VOUCHER_TYPE = "Credit Note (Goods Return)"
COMPANY_NAME = "Headphone Zone Pvt Ltd Chennai - (from 1-Apr-26)"

# ─── GST LEDGER MAP ────────────────────────────
GST_LEDGER_MAP = {
    5:  ("Output CGST @2.5%",  "Output SGST @2.5%",  "Output IGST @5%"),
    12: ("Output CGST @6%",    "Output SGST @6%",    "Output IGST @12%"),
    18: ("Output CGST @9%",    "Output SGST @9%",    "Output IGST @18%"),
    28: ("Output CGST @14%",   "Output SGST @14%",   "Output IGST @28%"),
}

# ─── SALES LEDGER MAP ──────────────────────────
SALES_LEDGER_MAP = {
    5:  "Local GST5.0%",
    12: "Local GST12.0%",
    18: "Local GST18.0%",
    28: "Local GST28.0%",
}

# ─────────────────────────────────────────────
# UNICOMMERCE AUTH & API (all use internal IP)
# ─────────────────────────────────────────────

def get_public_ip() -> str:
    try:
        return requests.get('https://api.ipify.org', timeout=5).text
    except:
        return "unknown"

def get_uc_token(password: str, retries: int = 3) -> str:
    url = f"{UC_BASE}/oauth/token"
    params = {
        "grant_type": "password",
        "client_id": "my-trusted-client",
        "username": UC_USERNAME,
        "password": password,
    }
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, timeout=30)
            if r.status_code == 403:
                raise ValueError(f"403 Forbidden on token: {r.text[:200]}")
            r.raise_for_status()
            token = r.json().get("access_token", "").strip()
            if not token:
                raise ValueError("Empty token received")
            return token
        except requests.exceptions.RequestException:
            if attempt == retries - 1:
                raise
            wait = 2 ** attempt
            time.sleep(wait)
    raise RuntimeError("Failed to get token after retries")

def uc_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Facility": UC_FACILITY,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
    }

def search_sale_order(display_order_code: str, token: str, retries: int = 3) -> dict | None:
    # Use internal IP for search as well
    url = f"{UC_BASE}/services/rest/v1/oms/saleOrder/search"
    payload = {"displayOrderCode": display_order_code}
    for attempt in range(retries):
        try:
            r = requests.post(url, json=payload, headers=uc_headers(token), timeout=15)
            if r.status_code == 403:
                public_ip = get_public_ip()
                raise ValueError(f"403 Forbidden from IP {public_ip}. Please whitelist this IP or use internal network. Response: {r.text[:300]}")
            r.raise_for_status()
            data = r.json()
            if not data.get("successful"):
                raise ValueError(f"UC search failed: {data.get('errors')}")
            elements = data.get("elements", [])
            return elements[0] if elements else None
        except requests.exceptions.RequestException:
            if attempt == retries - 1:
                raise
            wait = 2 ** attempt
            time.sleep(wait)
    return None

def get_sale_order(sale_order_code: str, token: str, retries: int = 3) -> dict:
    url = f"{UC_BASE}/oms/saleorder/get"
    payload = {
        "code": sale_order_code,
        "facilityCodes": [UC_FACILITY],
        "paymentDetailRequired": True,
    }
    for attempt in range(retries):
        try:
            r = requests.post(url, json=payload, headers=uc_headers(token), timeout=20)
            if r.status_code == 403:
                public_ip = get_public_ip()
                raise ValueError(f"403 Forbidden from IP {public_ip}. Please whitelist this IP or use internal network. Response: {r.text[:300]}")
            r.raise_for_status()
            data = r.json()
            if not data.get("successful"):
                raise ValueError(f"UC get sale order failed: {data.get('errors')}")
            return data["saleOrderDTO"]
        except requests.exceptions.RequestException:
            if attempt == retries - 1:
                raise
            wait = 2 ** attempt
            time.sleep(wait)
    raise RuntimeError("Failed to get sale order after retries")

def fetch_order_data(display_order_code: str, token: str) -> dict:
    element = search_sale_order(display_order_code, token)
    if not element:
        raise ValueError(f"No sale order found for {display_order_code}")
    sale_order_code = element["code"]
    so_dto = get_sale_order(sale_order_code, token)

    pkgs = so_dto.get("shippingPackages") or []
    if not pkgs:
        raise ValueError(f"No shipping packages in sale order {sale_order_code}")
    pkg = pkgs[0]

    invoice_display = pkg.get("invoiceDisplayCode") or pkg.get("invoiceCode", "")
    invoice_date_ms = pkg.get("invoiceDate")
    if invoice_date_ms:
        invoice_date_str = datetime.fromtimestamp(invoice_date_ms / 1000).strftime("%d-%m-%Y")
    else:
        invoice_date_str = ""

    customer_email = pkg.get("customer") or element.get("notificationEmail", "")
    cod_flag = so_dto.get("cod", False)

    sale_items = so_dto.get("saleOrderItems") or []
    pkg_code = pkg.get("code", "")
    pkg_items = [i for i in sale_items if i.get("shippingPackageCode") == pkg_code]
    if not pkg_items:
        pkg_items = sale_items

    billing_state = (so_dto.get("billingAddress") or {}).get("state", "TN")
    is_interstate = billing_state.upper() != "TN"

    # ─── Extract addresses ──────────────────────────
    billing_addr = so_dto.get("billingAddress") or {}
    billing_line = f"{billing_addr.get('addressLine1', '')} {billing_addr.get('addressLine2', '')} {billing_addr.get('city', '')}".strip()
    billing_pincode = billing_addr.get("pincode", "")
    billing_state_name = billing_addr.get("stateName", billing_addr.get("state", ""))

    addresses = so_dto.get("addresses") or []
    shipping_addr = None
    if len(addresses) >= 2:
        shipping_addr = addresses[1]
    else:
        shipping_addr = billing_addr
    shipping_line = f"{shipping_addr.get('addressLine1', '')} {shipping_addr.get('addressLine2', '')} {shipping_addr.get('city', '')}".strip()
    shipping_pincode = shipping_addr.get("pincode", "")
    shipping_state_name = shipping_addr.get("stateName", shipping_addr.get("state", ""))

    # ─── GST Registration ────────────────────────────
    customer_gstin = so_dto.get("customerGSTIN") or ""
    if customer_gstin and customer_gstin.strip():
        gst_reg_type = "Regular"
        gstin_value = customer_gstin.strip()
    else:
        gst_reg_type = "Unregistered/Consumer"
        gstin_value = ""

    return {
        "sale_order_code":  sale_order_code,
        "invoice_display":  invoice_display,
        "invoice_date_str": invoice_date_str,
        "customer_email":   customer_email,
        "cod":              cod_flag,
        "is_interstate":    is_interstate,
        "sale_items":       pkg_items,
        "billing_address":  billing_line,
        "billing_pincode":  billing_pincode,
        "billing_state":    billing_state_name,
        "shipping_address": shipping_line,
        "shipping_pincode": shipping_pincode,
        "shipping_state":   shipping_state_name,
        "place_of_supply":  billing_state_name,
        "gst_reg_type":     gst_reg_type,
        "gstin":            gstin_value,
        "raw_so":           so_dto,
    }

# ─────────────────────────────────────────────
# GOOGLE SHEET READER – BATCH READING
# ─────────────────────────────────────────────

def read_sheet_raw(sheet_url: str, gcp_creds: dict) -> pd.DataFrame:
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_info(gcp_creds, scopes=scopes)
    gc = gspread.authorize(creds)

    sheet_id_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url)
    if sheet_id_match:
        sheet_id = sheet_id_match.group(1)
    else:
        sheet_id = sheet_url.strip()

    spreadsheet = gc.open_by_key(sheet_id)

    try:
        ws = spreadsheet.worksheet("Form responses 1")
    except gspread.exceptions.WorksheetNotFound:
        st.error("Tab 'Form responses 1' not found. Available tabs: " + str([ws.title for ws in spreadsheet.worksheets()]))
        return pd.DataFrame()

    def get_all_rows_batched(worksheet, batch_size=5000):
        all_rows = []
        start_row = 1
        max_col_letter = 'Z'
        while True:
            end_row = start_row + batch_size - 1
            range_name = f"A{start_row}:{max_col_letter}{end_row}"
            try:
                batch = worksheet.get(range_name)
            except Exception as e:
                st.error(f"Error reading batch rows {start_row}-{end_row}: {e}")
                break
            if not batch:
                break
            all_rows.extend(batch)
            if len(batch) < batch_size:
                break
            start_row = end_row + 1
        return all_rows

    all_values = get_all_rows_batched(ws)
    if not all_values:
        return pd.DataFrame()

    headers = all_values[0]
    data_rows = all_values[1:]

    max_cols = len(headers)
    for i, row in enumerate(data_rows):
        if len(row) < max_cols:
            data_rows[i] = row + [""] * (max_cols - len(row))

    df = pd.DataFrame(data_rows, columns=headers)
    if df.empty:
        return df

    df.columns = df.columns.str.strip()

    # ─── Clean timestamp: remove commas, then parse ──
    df["Timestamp"] = df["Timestamp"].astype(str).str.strip().str.replace(",", " ")
    df["Timestamp"] = pd.to_datetime(df["Timestamp"], dayfirst=True, errors="coerce")

    return df

# ─────────────────────────────────────────────
# TALLY EXCEL BUILDER (with GST Registration fields)
# ─────────────────────────────────────────────

TALLY_COLUMNS = [
    "Voucher Date",
    "Voucher Type Name",
    "Voucher Number",
    "Reference No.",
    "Reference Date",
    "Voucher Narration",
    "Ledger Name",
    "Ledger Amount",
    "Ledger Amount Dr/Cr",
    "Bill Type of Ref",
    "Bill Name",
    "Bill Amount",
    "Bill Amount - Dr/Cr",
    "Buyer/Supplier - Address",
    "Buyer/Supplier - Pincode",
    "Buyer/Supplier - State",
    "Buyer/Supplier - GST Registration Type",
    "Buyer/Supplier - GSTIN/UIN",
    "Buyer/Supplier - Place of Supply",
    "Consignee - Address",
    "Consignee - Pincode",
    "Consignee - State",
    "Item Name",
    "Actual Quantity",
    "Billed Quantity",
    "Quantity UOM",
    "Item Rate",
    "Item Rate per",
    "Item Amount",
    "Item Allocations - Godown Name",
    "Item Allocations - Actual Quantity",
    "Item Allocations - Billed Quantity",
    "Item Allocations - UOM",
    "Item Allocations - Rate",
    "Item Allocations - Rate per",
    "Item Allocations - Amount",
    "Original Invoice No.",
    "Original Invoice - Date",
    "Nature of Original Sales",
    "Change Mode",
]

def _gst_slab(item: dict) -> int:
    cgst = item.get("centralGstPercentage", 0) or 0
    sgst = item.get("stateGstPercentage", 0) or 0
    igst = item.get("integratedGstPercentage", 0) or 0
    total = round(cgst + sgst + igst)
    for slab in [5, 12, 18, 28]:
        if abs(total - slab) <= 1:
            return slab
    return 18

def _is_interstate(item: dict) -> bool:
    return (item.get("integratedGstPercentage", 0) or 0) > 0

def build_excel(credit_notes: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounting Voucher"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(TALLY_COLUMNS)
    for col_idx, _ in enumerate(TALLY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 30

    data_font = Font(name="Arial", size=9)
    alt_fill = PatternFill("solid", start_color="EFF3FA")
    row_num = 2
    cn_count = 0

    for cn in credit_notes:
        cn_count += 1
        fill = alt_fill if cn_count % 2 == 0 else PatternFill()
        items = cn["items"]

        total_item_amt = 0.0
        total_cgst = 0.0
        total_sgst = 0.0
        total_igst = 0.0
        total_cod = cn.get("cod_charges", 0)

        for item in items:
            rate = round(item["unit_price"], 2)
            total_item_amt += (rate * 1)
            total_cgst += round(item.get("cgst_amt", 0), 2)
            total_sgst += round(item.get("sgst_amt", 0), 2)
            total_igst += round(item.get("igst_amt", 0), 2)

        slab = items[0].get("gst_slab", 18) if items else 18
        interstate = items[0].get("is_interstate", False) if items else False
        ledger_names = GST_LEDGER_MAP.get(slab, GST_LEDGER_MAP[18])

        if interstate:
            sales_ledger_name = f"Interstate GST {slab}.0%"
        else:
            sales_ledger_name = SALES_LEDGER_MAP.get(slab, f"Local GST{slab}.0%")

        col = {v: i for i, v in enumerate(TALLY_COLUMNS)}

        party_credit = round(total_item_amt + total_cgst + total_sgst + total_igst + total_cod, 2)

        # ── Party Ledger (Cr) – linked to original Order ID ──
        row1 = [""] * len(TALLY_COLUMNS)
        row1[col["Voucher Date"]] = cn["voucher_date"]
        row1[col["Voucher Type Name"]] = VOUCHER_TYPE
        row1[col["Voucher Number"]] = cn["order_no"]
        row1[col["Reference No."]] = cn["invoice_no"]
        row1[col["Reference Date"]] = cn["invoice_date"]
        row1[col["Voucher Narration"]] = cn["order_no"]
        row1[col["Ledger Name"]] = cn["customer_email"]
        row1[col["Ledger Amount"]] = party_credit
        row1[col["Ledger Amount Dr/Cr"]] = "Cr"
        row1[col["Bill Type of Ref"]] = "Agst Ref"
        row1[col["Bill Name"]] = cn["order_no"]
        row1[col["Bill Amount"]] = party_credit
        row1[col["Bill Amount - Dr/Cr"]] = "Cr"

        # Party details
        row1[col["Buyer/Supplier - Address"]] = cn.get("billing_address", "")
        row1[col["Buyer/Supplier - Pincode"]] = cn.get("billing_pincode", "")
        row1[col["Buyer/Supplier - State"]] = cn.get("billing_state", "")
        row1[col["Buyer/Supplier - GST Registration Type"]] = cn.get("gst_reg_type", "Unregistered/Consumer")
        row1[col["Buyer/Supplier - GSTIN/UIN"]] = cn.get("gstin", "")
        row1[col["Buyer/Supplier - Place of Supply"]] = cn.get("place_of_supply", cn.get("billing_state", ""))

        row1[col["Consignee - Address"]] = cn.get("shipping_address", "")
        row1[col["Consignee - Pincode"]] = cn.get("shipping_pincode", "")
        row1[col["Consignee - State"]] = cn.get("shipping_state", "")

        row1[col["Original Invoice No."]] = cn["invoice_no"]
        row1[col["Original Invoice - Date"]] = cn["invoice_date"]
        row1[col["Nature of Original Sales"]] = "B2C (Small)"
        row1[col["Change Mode"]] = "Item Invoice"
        ws.append(row1)
        _style_row(ws, row_num, data_font, fill, border)
        row_num += 1

        # ── Item rows (Dr) – no Bill allocation ──
        for item in items:
            rate = round(item["unit_price"], 2)
            item_amt = round(rate * 1, 2)
            row_item = [""] * len(TALLY_COLUMNS)
            row_item[col["Item Name"]] = item["item_name"]
            row_item[col["Actual Quantity"]] = 1
            row_item[col["Billed Quantity"]] = 1
            row_item[col["Quantity UOM"]] = "Nos"
            row_item[col["Item Rate"]] = rate
            row_item[col["Item Rate per"]] = "Nos"
            row_item[col["Item Amount"]] = item_amt
            row_item[col["Item Allocations - Godown Name"]] = GODOWN
            row_item[col["Item Allocations - Actual Quantity"]] = 1
            row_item[col["Item Allocations - Billed Quantity"]] = 1
            row_item[col["Item Allocations - UOM"]] = "Nos"
            row_item[col["Item Allocations - Rate"]] = rate
            row_item[col["Item Allocations - Rate per"]] = "Nos"
            row_item[col["Item Allocations - Amount"]] = item_amt
            row_item[col["Ledger Name"]] = sales_ledger_name
            row_item[col["Ledger Amount"]] = item_amt
            row_item[col["Ledger Amount Dr/Cr"]] = "Dr"
            ws.append(row_item)
            _style_row(ws, row_num, data_font, fill, border)
            row_num += 1

        # ── Tax & COD (Dr) – no Bill allocation ──
        def add_ledger_row(name, amount):
            nonlocal row_num
            row = [""] * len(TALLY_COLUMNS)
            row[col["Ledger Name"]] = name
            row[col["Ledger Amount"]] = amount
            row[col["Ledger Amount Dr/Cr"]] = "Dr"
            ws.append(row)
            _style_row(ws, row_num, data_font, fill, border)
            row_num += 1

        if total_cod and total_cod > 0:
            add_ledger_row("COD Charges", round(total_cod, 2))

        if interstate and total_igst > 0:
            add_ledger_row(ledger_names[2], round(total_igst, 2))
        else:
            if total_cgst > 0:
                add_ledger_row(ledger_names[0], round(total_cgst, 2))
            if total_sgst > 0:
                add_ledger_row(ledger_names[1], round(total_sgst, 2))

        # ─── No Round Off ledger ──────────────────

    # ─── Column widths ──────────────────────────────
    col_widths = {
        "Voucher Date": 13, "Voucher Type Name": 28, "Voucher Number": 14,
        "Reference No.": 22, "Reference Date": 13, "Voucher Narration": 14,
        "Ledger Name": 35, "Ledger Amount": 14, "Ledger Amount Dr/Cr": 8,
        "Bill Type of Ref": 12, "Bill Name": 16, "Bill Amount": 14, "Bill Amount - Dr/Cr": 10,
        "Buyer/Supplier - Address": 40, "Buyer/Supplier - Pincode": 14, "Buyer/Supplier - State": 20,
        "Buyer/Supplier - GST Registration Type": 20, "Buyer/Supplier - GSTIN/UIN": 20,
        "Buyer/Supplier - Place of Supply": 20,
        "Consignee - Address": 40, "Consignee - Pincode": 14, "Consignee - State": 20,
        "Item Name": 50, "Actual Quantity": 8, "Billed Quantity": 8,
        "Quantity UOM": 8, "Item Rate": 12, "Item Rate per": 10, "Item Amount": 12,
        "Item Allocations - Godown Name": 20,
        "Item Allocations - Actual Quantity": 8, "Item Allocations - Billed Quantity": 8,
        "Item Allocations - UOM": 8, "Item Allocations - Rate": 12,
        "Item Allocations - Rate per": 10, "Item Allocations - Amount": 12,
        "Original Invoice No.": 22, "Original Invoice - Date": 13,
        "Nature of Original Sales": 18, "Change Mode": 14,
    }
    for col_idx, col_name in enumerate(TALLY_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TALLY_COLUMNS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def _style_row(ws, row_num, font, fill, border):
    for cell in ws[row_num]:
        cell.font = font
        if fill.fill_type:
            cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

# ─────────────────────────────────────────────
# MATCH ITEMS – unique matching & net COD
# ─────────────────────────────────────────────

def match_items(sheet_rows: pd.DataFrame, uc_sale_items: list, is_interstate: bool) -> list:
    result = []
    used_uc_indices = set()

    uc_items = []
    for idx, uc_item in enumerate(uc_sale_items):
        sku = str(uc_item.get("itemSku", "")).lower()
        name = str(uc_item.get("itemName", "")).lower()
        seller_sku = str(uc_item.get("sellerSkuCode", "")).lower()
        search_text = f"{sku} {name} {seller_sku}"
        uc_items.append({
            "index": idx,
            "item": uc_item,
            "search_text": search_text,
            "words": set(search_text.split())
        })

    for _, srow in sheet_rows.iterrows():
        model_name = str(srow.get("Model Name", "")).strip().lower()
        model_words = {w for w in model_name.split() if len(w) > 3}

        best_match = None
        best_score = 0

        for uc in uc_items:
            if uc["index"] in used_uc_indices:
                continue
            overlap = len(model_words & uc["words"])
            if overlap > best_score:
                best_score = overlap
                best_match = uc

        if best_match is None:
            for uc in uc_items:
                if uc["index"] not in used_uc_indices:
                    best_match = uc
                    break

        if best_match is not None:
            used_uc_indices.add(best_match["index"])
            matched = best_match["item"]

            cgst_pct = float(matched.get("centralGstPercentage", 0) or 0)
            sgst_pct = float(matched.get("stateGstPercentage", 0) or 0)
            igst_pct = float(matched.get("integratedGstPercentage", 0) or 0)
            cgst_amt = float(matched.get("totalCentralGst", 0) or 0)
            sgst_amt = float(matched.get("totalStateGst", 0) or 0)
            igst_amt = float(matched.get("totalIntegratedGst", 0) or 0)
            taxable_val = float(matched.get("sellingPriceWithoutTaxesAndDiscount", 0) or 0)
            selling_price = float(matched.get("sellingPrice", 0) or 0)
            rate = taxable_val if taxable_val > 0 else selling_price

            # COD: inclusive → net
            cod_inclusive = float(matched.get("cashOnDeliveryCharges", 0) or 0)
            slab = _gst_slab(matched)
            if cod_inclusive > 0 and slab > 0:
                tax_rate = slab / 100.0
                cod_net = round(cod_inclusive / (1 + tax_rate), 2)
            else:
                cod_net = 0.0

            result.append({
                "item_name": str(srow.get("Model Name", matched.get("itemName", ""))),
                "unit_price": rate,
                "selling_price": selling_price,
                "cgst_pct": cgst_pct,
                "sgst_pct": sgst_pct,
                "igst_pct": igst_pct,
                "cgst_amt": cgst_amt,
                "sgst_amt": sgst_amt,
                "igst_amt": igst_amt,
                "gst_slab": slab,
                "is_interstate": is_interstate,
                "cod_charges": cod_net,
            })

    return result

# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(page_title="HPZ Credit Note Generator", page_icon="📋", layout="wide")
st.markdown("""
    <style>
    .main { background: #f8f9fb; }
    h1 { color: #1F4E79; }
    .stButton > button {
        background: #1F4E79; color: white;
        border-radius: 6px; border: none;
        font-weight: 600; padding: 0.5rem 2rem;
    }
    .stButton > button:hover { background: #2e6da4; }
    </style>
""", unsafe_allow_html=True)

st.title("📋 HPZ Credit Note Generator")
st.caption("Goods Return · Pulls data from RMS Google Sheet + Unicommerce · Exports Tally-ready Excel")

col1, col2 = st.columns(2)
with col1:
    from_date = st.date_input("From Date", value=date.today() - timedelta(days=30))
with col2:
    to_date = st.date_input("To Date", value=date.today())

if from_date > to_date:
    st.error("From Date must be before To Date.")
    st.stop()

st.divider()
generate_btn = st.button("🚀 Generate Credit Notes", width='stretch')

if generate_btn:
    # ─── Read secrets ──────────────────────────────
    try:
        uc_password = st.secrets["UC_PASSWORD"]
        sheet_url = st.secrets["SHEET_ID"]
        gcp_creds_json = st.secrets["GCP_SA_JSON"]
    except KeyError as e:
        st.error(f"Missing secret: {e}. Please add it to your .streamlit/secrets.toml or Streamlit Cloud secrets.")
        st.stop()

    try:
        gcp_creds = json.loads(gcp_creds_json)
    except json.JSONDecodeError:
        st.error("Invalid GCP service account JSON in secrets.")
        st.stop()

    progress = st.progress(0, text="Starting…")
    log_box = st.empty()
    logs = []

    def log(msg, error=False):
        icon = "❌" if error else "✅"
        logs.append(f"{icon} {msg}")
        log_box.markdown("\n\n".join(logs))

    try:
        progress.progress(5, text="Reading Google Sheet…")
        raw_df = read_sheet_raw(sheet_url, gcp_creds)

        if raw_df.empty:
            st.error("No data returned from the sheet. Check the tab name and permissions.")
            st.stop()

        from_dt = datetime.combine(from_date, datetime.min.time())
        to_dt   = datetime.combine(to_date,   datetime.max.time())
        df = raw_df[(raw_df["Timestamp"] >= from_dt) & (raw_df["Timestamp"] <= to_dt)]
        df = df[df["Return Status"].str.strip().str.lower() == "return good"]

        progress.progress(10, text="Filtering done")
        if df.empty:
            st.warning("No 'Return Good' rows found in the selected date range.")
            st.stop()

        log(f"Found {len(df)} rows ({df['Order No'].nunique()} unique orders)")

        progress.progress(15, text="Authenticating with Unicommerce…")
        token = get_uc_token(uc_password)
        log("Unicommerce auth OK")

        unique_orders = df["Order No"].astype(str).unique()
        credit_notes = []
        errors = []
        today_str = date.today().strftime("%d-%m-%Y")

        for idx, order_no in enumerate(unique_orders):
            pct = 15 + int(75 * (idx / len(unique_orders)))
            progress.progress(pct, text=f"Fetching order {order_no} ({idx+1}/{len(unique_orders)})…")

            try:
                display_code = f"#{order_no}"
                order_data = fetch_order_data(display_code, token)
                order_rows = df[df["Order No"].astype(str) == str(order_no)].copy()

                invoice_no = order_data["invoice_display"] or str(order_rows.iloc[0].get("Invoice No", order_no))
                inv_date_str = order_data["invoice_date_str"] or str(order_rows.iloc[0].get("Invoice date", "")).strip()

                first_row = order_rows.iloc[0]
                if pd.notna(first_row["Timestamp"]):
                    voucher_date = first_row["Timestamp"].strftime("%d-%m-%Y")
                else:
                    voucher_date = today_str

                matched_items = match_items(order_rows, order_data["sale_items"], order_data["is_interstate"])
                if not matched_items:
                    errors.append(f"Order {order_no}: Could not match any items.")
                    continue

                cod_charges = sum(it.get("cod_charges", 0) for it in matched_items) if order_data["cod"] else 0

                credit_notes.append({
                    "order_no": order_no,
                    "invoice_no": invoice_no,
                    "invoice_date": inv_date_str,
                    "voucher_date": voucher_date,
                    "customer_email": order_data["customer_email"],
                    "cod_charges": cod_charges,
                    "items": matched_items,
                    "billing_address": order_data.get("billing_address", ""),
                    "billing_pincode": order_data.get("billing_pincode", ""),
                    "billing_state": order_data.get("billing_state", ""),
                    "shipping_address": order_data.get("shipping_address", ""),
                    "shipping_pincode": order_data.get("shipping_pincode", ""),
                    "shipping_state": order_data.get("shipping_state", ""),
                    "place_of_supply": order_data.get("place_of_supply", order_data.get("billing_state", "")),
                    "gst_reg_type": order_data.get("gst_reg_type", "Unregistered/Consumer"),
                    "gstin": order_data.get("gstin", ""),
                })
                log(f"Order #{order_no} → Invoice {invoice_no} · {len(matched_items)} item(s)")

            except Exception as e:
                errors.append(f"Order {order_no}: {e}")
                log(f"Order #{order_no} FAILED: {e}", error=True)

        progress.progress(92, text="Building Excel file…")
        if credit_notes:
            excel_bytes = build_excel(credit_notes)
            log(f"Excel built: {len(credit_notes)} credit note(s)")
        else:
            st.error("No credit notes could be built. See errors above.")
            st.stop()

        progress.progress(100, text="Done!")

        # ─── Clear real-time log box ────────────────────
        log_box.empty()

        # ─── TOP: Success + Download button ──────────────
        st.success(f"✅ {len(credit_notes)} credit notes ready for download!")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.write(f"**{len(unique_orders)}** orders processed, **{len(errors)}** errors.")
        with col2:
            filename = f"CreditNotes_{from_date.strftime('%d%b%Y')}_to_{to_date.strftime('%d%b%Y')}.xlsx"
            st.download_button(
                label="⬇️  Download Excel",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch',
            )

        # ─── Processing logs (now static) ──────────────────
        if logs:
            st.subheader("📝 Processing Logs")
            st.code("\n".join(logs), language="text")

        # ─── Metrics ──────────────────────────────────────
        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.metric("Orders processed", len(unique_orders))
        c2.metric("Credit notes ready", len(credit_notes))
        c3.metric("Errors", len(errors))

        # ─── Errors expander ────────────────────────────
        if errors:
            with st.expander("⚠️ Orders with errors"):
                for e in errors:
                    st.error(e)

        # ─── Preview Table ─────────────────────────────
        st.subheader("📊 Preview (Item-level)")
        preview_rows = []
        for cn in credit_notes:
            for it in cn["items"]:
                preview_rows.append({
                    "Order No": cn["order_no"],
                    "Invoice No": cn["invoice_no"],
                    "Invoice Date": cn["invoice_date"],
                    "Customer": cn["customer_email"],
                    "Item": it["item_name"][:60],
                    "Rate": it["unit_price"],
                    "GST %": it["gst_slab"],
                    "Type": "IGST" if it["is_interstate"] else "CGST+SGST",
                })
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, height=300)

        st.success(f"Ready! Import **{filename}** into TallyPrime via Data → Import → Excel.")

    except Exception as ex:
        st.error(f"Unexpected error: {ex}")
        with st.expander("Traceback"):
            st.code(traceback.format_exc())